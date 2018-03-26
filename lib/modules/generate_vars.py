#!/usr/bin/python
import openpyxl
from ansible.module_utils.basic import *
import yaml
import os
def read_xls_dict(input_file, dest):
    "Read the XLS file and return as Ansible facts"
    result = {"ansible_facts":{}}
    spreadsheet = {}
    try:
        wb = openpyxl.load_workbook(input_file)
        spreadsheet = {}
        
        if dest in wb.sheetnames:
          current_sheet = wb.get_sheet_by_name(dest)
          dict_keys = []
          for c in range(2,current_sheet.max_column + 1): #first column is always hostnames
              dict_keys.append(current_sheet.cell(row=1,column=c).value)
          for r in range (2,current_sheet.max_row + 1):
              temp_dict = {}
              for c in range(2,current_sheet.max_column + 1):
                  temp_dict[dict_keys[c-2]] = current_sheet.cell(row=r,column=c).value
              spreadsheet[current_sheet.cell(row=r,column=1).value]=temp_dict
          
          #create host_vars/group_vars folder if it does not exist
          try: 
              os.makedirs(dest)
          except OSError:
              if not os.path.isdir(dest):
                  raise
          
          
          for k in spreadsheet.keys():
              with open(dest+'/'+str(k),'w') as yaml_file:
                  out = yaml.safe_dump(spreadsheet[k], default_flow_style=False, explicit_start=True)
                  out = re.sub(r"(.+:\s)'\[(.*)\]'", r'\1[\2]', out, flags=re.IGNORECASE)
                  yaml_file.write(out)
    except IOError:
        return (1, "IOError on input file:%s" % input_file)
    return (0, "")


def main():

    fields = {
        "src": {"required": True, "type": "str" }
    }
    module = AnsibleModule(argument_spec=fields)
    code, response = read_xls_dict(module.params["src"], "host_vars")
    code, response = read_xls_dict(module.params["src"], "group_vars")
    if code == 1:
        module.fail_json(msg=response)
    else:
        module.exit_json(changed=True, meta=module.params)

    return code



main()
#