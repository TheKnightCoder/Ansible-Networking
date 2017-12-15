#!/usr/bin/python
import re
import json
import fasteners
from pprint import pprint
from ansible.module_utils.basic import *
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from collections import OrderedDict

def WriteDictToXl(file_path, hostname, data):
  
  bold = Font(bold=True)
  file_exists = os.path.isfile(file_path)
  if file_exists:
    wb = load_workbook(file_path)
    ws = wb.create_sheet(hostname)
  else:
    wb = Workbook()
    ws = wb.active
    ws.title = hostname
    
  for i in range(0,len(data[0])): #adding column headers
    cell=ws.cell(row=1, column=i+1, value=data[0].keys()[i])
    cell.font = bold
 
  greenFill = PatternFill(fill_type='solid', start_color='c6efce', end_color='c6efce')
  for item in data:
    cur_row = ws.max_row+1
    for i in range(0,len(item)):
      if(isinstance(item.values()[i], list)): #if values in col is a list convert to a string
        item.values()[i]= str('\n'.join(item.values()[i]))
      cell = ws.cell(row=cur_row, column=i+1, value=str(item.values()[i]))
      if item.values()[i] == "on":
        cell.fill = greenFill
    
  wb.save(file_path)
  return

def main():
  fields = {
    "poe": {"required": True, "type": "dict" },
    "cdp": {"required": True, "type": "dict" },
    "dest": {"required": True, "type": "str" },
    "hostname": {"required": True, "type": "str" }
  }
  
  module = AnsibleModule(argument_spec=fields)
  poeList = []
  cdpList = []
  poeList = module.params['poe']['response']
  cdpList = module.params['cdp']['response']
  
  #merging the two tables into preferred output
  merged = []

  #Matching poeList interfaces with cdpList interfaces
  if poeList:
    for poeInterface in poeList:
      found = False
      if cdpList:
        for cdpNeighbor in cdpList:
          poe = re.search('([a-zA-z]+)(\d\S*)', poeInterface['interface']).group(1)
          poeName = poe.group(1)
          poeInt = poe.group(2)
          
          cdp = cdpNeighbor['local_interface'].split(" ")
          cdpName = cdp[0]
          cdpInt = cdp[1]
          
          if (cdpName in poeName) or (poeName in cdpName):
            if cdpInt == poeInt:
              found = True
              merged.append(OrderedDict([
                ("Interface", poeInterface['interface']),
                ("POE Status", poeInterface['operation']),
                ("AP", cdpNeighbor['neighbor'])
              ]))
              cdpList[:] = [d for d in cdpList if d.get('local_interface') != cdpNeighbor['local_interface']] #remove current cdp that has a match from cdp list
              break
      if not found: #add poe that has no match
        merged.append(OrderedDict([
            ("Interface", poeInterface['interface']),
            ("POE Status", poeInterface['operation']),
            ("AP", "")
        ]))
        
  for cdpNeighbor in cdpList: #add cdp that has no match
    found = False
    merged.append(OrderedDict([
      ("Interface", cdpNeighbor['local_interface']),
      ("POE Status", 'N/A'),
      ("AP", cdpNeighbor['neighbor'])
    ]))

  try:
    with fasteners.InterProcessLock('/tmp/ansible_lock_file'):
      WriteDictToXl(module.params['dest'], module.params['hostname'], merged)
  except IOError as (errno, strerror):
    print("I/O error({0}): {1}".format(errno, strerror))
    module.fail_json(msg="I/O error({0}): {1}".format(errno, strerror))

  module.exit_json(changed=True, meta=module.params)

if __name__ == '__main__':  
  main()
