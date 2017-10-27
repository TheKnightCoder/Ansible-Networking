#!/usr/bin/python

import json
import fasteners
from pprint import pprint
from ansible.module_utils.basic import *
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from collections import OrderedDict

def WriteDictToXl(file_path, hostname, data):
  
  bold = Font(bold=True)
  file_exists = os.path.isfile(file_path)
  if file_exists:
    wb = load_workbook(file_path)
    ws = wb.create_sheet(hostname)
  else:
    wb = Workbook()
    ws = wb.active
    ws.title = hostname
    
  for i in range(0,len(data[0])): #adding column headers
    cell=ws.cell(row=1, column=i+1, value=data[0].keys()[i])
    cell.font = bold
 
  greenFill = PatternFill(fill_type='solid', start_color='c6efce', end_color='c6efce')
  for item in data:
    cur_row = ws.max_row+1
    for i in range(0,len(item)):
      if(isinstance(item.values()[i], list)): #if values in col is a list convert to a string
        item.values()[i]= str('\n'.join(item.values()[i]))
      cell = ws.cell(row=cur_row, column=i+1, value=str(item.values()[i]))
      if item.values()[i] == "on":
        cell.fill = greenFill
    
  wb.save(file_path)
  return

def main():
  fields = {
    "poe": {"required": True, "type": "dict" },
    "cdp": {"required": True, "type": "dict" },
    "dest": {"required": True, "type": "str" },
    "hostname": {"required": True, "type": "str" }
  }
  
  module = AnsibleModule(argument_spec=fields)
  
  poe = module.params['poe']['response']
  cdp = module.params['cdp']['response']
  
  #merging the two tables into preferred output
  merged = []
  for c in cdp:
    c['local_interface'] = c['local_interface'].split(" ")[1]
    print c['local_interface']
  for p in poe:
    found = False
    for c in cdp:
      if c['local_interface'] in p['interface']:
        found = True
        merged.append(OrderedDict([
          ("Interface", p['interface']),
          ("POE Status", p['operation']),
          ("AP", c['neighbor'])
        ]))
    if not found:
      merged.append(OrderedDict([
          ("Interface", p['interface']),
          ("POE Status", p['operation']),
          ("AP", "")
      ]))

  try:
    with fasteners.InterProcessLock('/tmp/ansible_lock_file'):
      WriteDictToXl(module.params['dest'], module.params['hostname'], merged)
  except IOError as (errno, strerror):
    print("I/O error({0}): {1}".format(errno, strerror))
    module.fail_json(msg="I/O error({0}): {1}".format(errno, strerror))

  module.exit_json(changed=True, meta=module.params)

if __name__ == '__main__':  
  main()
