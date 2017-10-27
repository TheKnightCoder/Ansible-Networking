#!/usr/bin/python

import json
import fasteners
from pprint import pprint
from ansible.module_utils.basic import *
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook

#def WriteDictToCSV(csv_file,dict_data):
#import csv
#  try:
#    file_exists = os.path.isfile(csv_file)
#    with open(csv_file, 'ab') as csvfile:
#      writer = csv.DictWriter(csvfile, dict_data.keys())
#      if not file_exists:
#        writer.writeheader()
#      #for data in dict_data:
#      writer.writerow(dict_data)
#  except IOError as (errno, strerror):
#    print("I/O error({0}): {1}".format(errno, strerror))
#    module.fail_json(msg="I/O error({0}): {1}".format(errno, strerror))
#  return   

def WriteDictToXl(csv_file,dict_data):

  file_exists = os.path.isfile(csv_file)
  if file_exists:
    wb = load_workbook(csv_file)
    ws = wb.active
  else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Facts"
    for i in range(0,len(dict_data)): #adding column headers
      ws.cell(row=1, column=i+1, value=dict_data.keys()[i])
  
  cur_row = ws.max_row+1
  for i in range(0,len(dict_data)):
    if(isinstance(dict_data.values()[i], list)): #if a column/value has multiple values and is an array then turn into string e.g. dict_data['interface']= [1,2,3] ==> 1 /n 2 /n 3
      dict_data.values()[i]= str('\n'.join(dict_data.values()[i]))
    ws.cell(row=cur_row, column=i+1, value=str(dict_data.values()[i]))
    
  wb.save(csv_file)
  return

def main():
  fields = {
    "content": {"required": True, "type": "dict" },
    "dest": {"required": True, "type": "str" },
  }
  
  module = AnsibleModule(argument_spec=fields)
  
  facts = module.params['content']['ansible_facts']['napalm_facts']
  try:
    with fasteners.InterProcessLock('/tmp/ansible_lock_file'):
      WriteDictToXl(module.params['dest'], facts)
  except IOError as (errno, strerror):
    print("I/O error({0}): {1}".format(errno, strerror))
    module.fail_json(msg="I/O error({0}): {1}".format(errno, strerror))

  module.exit_json(changed=True, meta=module.params)

if __name__ == '__main__':  
  main()
